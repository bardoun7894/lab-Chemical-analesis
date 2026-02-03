"""
Excel Import Service

Imports data from Excel files (Master.xlsb format) into the database.
Supports importing chemical analyses, pipes, stages, and mechanical tests.
"""

from datetime import date, datetime
from typing import Dict, List, Tuple, Optional
import os

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from app import db
from app.models.chemical import ChemicalAnalysis, Furnace
from app.models.pipe import Pipe, PipeStage
from app.models.mechanical import MechanicalTest
from app.services.ladle_utils import generate_ladle_id


class ExcelImportError(Exception):
    """Custom exception for import errors"""
    pass


class ExcelImporter:
    """
    Handles importing data from Excel files.
    """

    def __init__(self, file_path: str):
        """
        Initialize importer with file path.

        Args:
            file_path: Path to Excel file
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel import")

        self.file_path = file_path
        self.workbook = None
        self.errors = []
        self.warnings = []
        self.imported_counts = {
            'chemical': 0,
            'pipes': 0,
            'stages': 0,
            'mechanical': 0
        }

    def load_file(self):
        """Load the Excel file"""
        if not os.path.exists(self.file_path):
            raise ExcelImportError(f"File not found: {self.file_path}")

        try:
            # Note: openpyxl can read .xlsx but not .xlsb
            # For .xlsb files, user needs to convert to .xlsx first
            self.workbook = load_workbook(self.file_path, data_only=True)
        except Exception as e:
            raise ExcelImportError(f"Error loading file: {str(e)}")

    def get_sheet_names(self) -> List[str]:
        """Get list of sheet names"""
        if not self.workbook:
            self.load_file()
        return self.workbook.sheetnames

    def preview_sheet(self, sheet_name: str, max_rows: int = 10) -> List[List]:
        """
        Preview data from a sheet.

        Args:
            sheet_name: Name of sheet to preview
            max_rows: Maximum rows to return

        Returns:
            List of rows (each row is a list of cell values)
        """
        if not self.workbook:
            self.load_file()

        if sheet_name not in self.workbook.sheetnames:
            raise ExcelImportError(f"Sheet not found: {sheet_name}")

        sheet = self.workbook[sheet_name]
        rows = []

        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i >= max_rows:
                break
            rows.append(list(row))

        return rows

    def import_chemical_analysis(self, sheet_name: str = 'Lab Chemical Analysis Table',
                                  start_row: int = 14,
                                  column_mapping: Dict = None) -> int:
        """
        Import chemical analysis data from sheet.

        Expected columns (default mapping):
        - A: Furnace Code
        - B: Ladle Number
        - C: Day
        - D: Month
        - E: Year
        - F-O: Element values (C, Si, Mg, Cu, Cr, S, Mn, P, Pb, Al)
        - P-R: Calculated values (CE, MnE, MgE)
        - S: Decision
        - T: Reason/Notes

        Args:
            sheet_name: Name of sheet containing chemical data
            start_row: Row number where data starts (1-indexed)
            column_mapping: Optional custom column mapping

        Returns:
            Number of records imported
        """
        if not self.workbook:
            self.load_file()

        if sheet_name not in self.workbook.sheetnames:
            raise ExcelImportError(f"Sheet not found: {sheet_name}")

        sheet = self.workbook[sheet_name]

        # Default column mapping (0-indexed)
        default_mapping = {
            'furnace_code': 0,
            'ladle_no': 1,
            'day': 2,
            'month': 3,
            'year': 4,
            'carbon': 5,
            'silicon': 6,
            'magnesium': 7,
            'copper': 8,
            'chromium': 9,
            'sulfur': 10,
            'manganese': 11,
            'phosphorus': 12,
            'lead': 13,
            'aluminum': 14,
            'carbon_equivalent': 15,
            'manganese_equivalent': 16,
            'magnesium_equivalent': 17,
            'decision': 18,
            'notes': 19
        }

        mapping = column_mapping or default_mapping
        count = 0

        # Get furnace mapping
        furnaces = {f.furnace_code: f.id for f in Furnace.query.all()}

        for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            try:
                # Skip empty rows
                if not row or not row[mapping['ladle_no']]:
                    continue

                # Parse data
                furnace_code = str(row[mapping['furnace_code']]).strip() if row[mapping['furnace_code']] else None
                ladle_no = int(row[mapping['ladle_no']]) if row[mapping['ladle_no']] else None
                day = int(row[mapping['day']]) if row[mapping['day']] else None
                month = int(row[mapping['month']]) if row[mapping['month']] else None
                year = int(row[mapping['year']]) if row[mapping['year']] else None

                if not all([ladle_no, day, month, year]):
                    self.warnings.append(f"Row {row_num}: Missing required date/ladle info")
                    continue

                # Create test date
                try:
                    test_date = date(year, month, day)
                except ValueError:
                    self.errors.append(f"Row {row_num}: Invalid date {day}/{month}/{year}")
                    continue

                # Generate ladle_id
                ladle_id = generate_ladle_id(ladle_no, test_date)

                # Check for duplicate
                existing = ChemicalAnalysis.query.filter_by(ladle_id=ladle_id).first()
                if existing:
                    self.warnings.append(f"Row {row_num}: Duplicate ladle_id {ladle_id}, skipping")
                    continue

                # Create analysis record
                analysis = ChemicalAnalysis(
                    test_date=test_date,
                    furnace_id=furnaces.get(furnace_code),
                    ladle_no=ladle_no,
                    day=day,
                    month=month,
                    year=year,
                    ladle_id=ladle_id,
                    carbon=safe_float(row[mapping['carbon']]),
                    silicon=safe_float(row[mapping['silicon']]),
                    magnesium=safe_float(row[mapping['magnesium']]),
                    copper=safe_float(row[mapping['copper']]),
                    chromium=safe_float(row[mapping['chromium']]),
                    sulfur=safe_float(row[mapping['sulfur']]),
                    manganese=safe_float(row[mapping['manganese']]),
                    phosphorus=safe_float(row[mapping['phosphorus']]),
                    lead=safe_float(row[mapping['lead']]),
                    aluminum=safe_float(row[mapping['aluminum']]),
                    carbon_equivalent=safe_float(row[mapping['carbon_equivalent']]),
                    manganese_equivalent=safe_float(row[mapping['manganese_equivalent']]),
                    magnesium_equivalent=safe_float(row[mapping['magnesium_equivalent']]),
                    decision=str(row[mapping['decision']]).strip().upper() if row[mapping['decision']] else None,
                    notes=str(row[mapping['notes']]).strip() if row[mapping['notes']] else None
                )

                db.session.add(analysis)
                count += 1

                # Commit in batches
                if count % 100 == 0:
                    db.session.commit()

            except Exception as e:
                self.errors.append(f"Row {row_num}: {str(e)}")
                continue

        db.session.commit()
        self.imported_counts['chemical'] = count
        return count

    def import_pipes_and_stages(self, sheet_name: str = 'Stages Tables',
                                start_row: int = 2,
                                column_mapping: Dict = None) -> Tuple[int, int]:
        """
        Import pipes and their production stages.

        Args:
            sheet_name: Name of sheet containing stage data
            start_row: Row number where data starts
            column_mapping: Optional custom column mapping

        Returns:
            Tuple of (pipes imported, stages imported)
        """
        if not self.workbook:
            self.load_file()

        if sheet_name not in self.workbook.sheetnames:
            raise ExcelImportError(f"Sheet not found: {sheet_name}")

        sheet = self.workbook[sheet_name]

        # Default column mapping for stages sheet
        default_mapping = {
            'production_date': 0,
            'shift': 1,
            'no_code': 2,
            'ladle_id': 3,
            'diameter': 4,
            'pipe_type': 5,
            'actual_weight': 6,
            # Stages start from column 7
            'stage_start_col': 7
        }

        mapping = column_mapping or default_mapping

        pipe_count = 0
        stage_count = 0

        stage_names = ['CCM', 'Annealing', 'Zinc', 'Cutting', 'Hydrotest', 'Cement', 'Coating', 'Finish']

        for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            try:
                if not row or not row[mapping['no_code']]:
                    continue

                no_code = str(row[mapping['no_code']]).strip()

                # Check for duplicate pipe
                existing = Pipe.query.filter_by(no_code=no_code).first()
                if existing:
                    self.warnings.append(f"Row {row_num}: Duplicate pipe {no_code}, skipping")
                    continue

                # Parse production date
                prod_date = row[mapping['production_date']]
                if isinstance(prod_date, datetime):
                    prod_date = prod_date.date()
                elif isinstance(prod_date, str):
                    try:
                        prod_date = datetime.strptime(prod_date, '%Y-%m-%d').date()
                    except:
                        prod_date = None

                # Create pipe
                pipe = Pipe(
                    production_date=prod_date,
                    shift=safe_int(row[mapping['shift']]),
                    no_code=no_code,
                    ladle_id=str(row[mapping['ladle_id']]).strip() if row[mapping['ladle_id']] else None,
                    diameter=safe_int(row[mapping['diameter']]),
                    pipe_type=str(row[mapping['pipe_type']]).strip() if row[mapping['pipe_type']] else None,
                    actual_weight=safe_float(row[mapping['actual_weight']])
                )

                db.session.add(pipe)
                db.session.flush()  # Get pipe.id
                pipe_count += 1

                # Import stages (each stage may have decision in subsequent columns)
                stage_col = mapping['stage_start_col']
                for stage_name in stage_names:
                    if stage_col < len(row):
                        decision = row[stage_col]
                        if decision:
                            stage = PipeStage(
                                pipe_id=pipe.id,
                                stage_name=stage_name,
                                decision=str(decision).strip().upper() if decision else None
                            )
                            db.session.add(stage)
                            stage_count += 1
                        stage_col += 1

                if pipe_count % 100 == 0:
                    db.session.commit()

            except Exception as e:
                self.errors.append(f"Row {row_num}: {str(e)}")
                continue

        db.session.commit()
        self.imported_counts['pipes'] = pipe_count
        self.imported_counts['stages'] = stage_count
        return pipe_count, stage_count

    def import_mechanical_tests(self, sheet_name: str = 'Lab Mech. Tables',
                                start_row: int = 2,
                                column_mapping: Dict = None) -> int:
        """
        Import mechanical test data.

        Args:
            sheet_name: Name of sheet containing mechanical test data
            start_row: Row number where data starts
            column_mapping: Optional custom column mapping

        Returns:
            Number of records imported
        """
        if not self.workbook:
            self.load_file()

        if sheet_name not in self.workbook.sheetnames:
            raise ExcelImportError(f"Sheet not found: {sheet_name}")

        sheet = self.workbook[sheet_name]

        # Default column mapping
        default_mapping = {
            'test_date': 0,
            'diameter': 1,
            'code': 2,
            'ladle_id': 3,
            'sample_thickness': 4,
            'd1': 5,
            'd2': 6,
            'd3': 7,
            'force_kgf': 8,
            'tensile_strength': 9,
            'elongation': 10,
            'nodularity_percent': 11,
            'hardness': 12,
            'decision': 13,
            'comments': 14
        }

        mapping = column_mapping or default_mapping
        count = 0

        for row_num, row in enumerate(sheet.iter_rows(min_row=start_row, values_only=True), start=start_row):
            try:
                if not row or not row[mapping['code']]:
                    continue

                # Parse test date
                test_date = row[mapping['test_date']]
                if isinstance(test_date, datetime):
                    test_date = test_date.date()
                elif isinstance(test_date, str):
                    try:
                        test_date = datetime.strptime(test_date, '%Y-%m-%d').date()
                    except:
                        test_date = date.today()

                mech_test = MechanicalTest(
                    test_date=test_date,
                    diameter=safe_int(row[mapping['diameter']]),
                    code=str(row[mapping['code']]).strip() if row[mapping['code']] else None,
                    ladle_id=str(row[mapping['ladle_id']]).strip() if row[mapping['ladle_id']] else None,
                    sample_thickness=safe_float(row[mapping['sample_thickness']]),
                    d1=safe_float(row[mapping['d1']]),
                    d2=safe_float(row[mapping['d2']]),
                    d3=safe_float(row[mapping['d3']]),
                    force_kgf=safe_float(row[mapping['force_kgf']]),
                    tensile_strength=safe_float(row[mapping['tensile_strength']]),
                    elongation=safe_float(row[mapping['elongation']]),
                    nodularity_percent=safe_float(row[mapping['nodularity_percent']]),
                    hardness=safe_float(row[mapping['hardness']]),
                    decision=str(row[mapping['decision']]).strip().upper() if row[mapping['decision']] else None,
                    comments=str(row[mapping['comments']]).strip() if row[mapping['comments']] else None
                )

                db.session.add(mech_test)
                count += 1

                if count % 100 == 0:
                    db.session.commit()

            except Exception as e:
                self.errors.append(f"Row {row_num}: {str(e)}")
                continue

        db.session.commit()
        self.imported_counts['mechanical'] = count
        return count

    def get_summary(self) -> Dict:
        """Get import summary"""
        return {
            'imported': self.imported_counts,
            'errors': self.errors,
            'warnings': self.warnings,
            'error_count': len(self.errors),
            'warning_count': len(self.warnings)
        }

    def close(self):
        """Close workbook"""
        if self.workbook:
            self.workbook.close()


def safe_float(value) -> Optional[float]:
    """Safely convert value to float"""
    if value is None:
        return None
    try:
        return float(value)
    except (ValueError, TypeError):
        return None


def safe_int(value) -> Optional[int]:
    """Safely convert value to int"""
    if value is None:
        return None
    try:
        return int(float(value))
    except (ValueError, TypeError):
        return None
